from fastapi import FastAPI, Depends, HTTPException, Query, UploadFile, File
from fastapi.middleware.cors import CORSMiddleware
from fastapi.staticfiles import StaticFiles
from fastapi.responses import FileResponse, StreamingResponse
from sqlalchemy.orm import Session
from pydantic import BaseModel
from typing import Optional, List
from datetime import date as date_type, timedelta
import models
import crud
import pathlib
import os
import io
from database import engine, get_db, Base

Base.metadata.create_all(bind=engine)

app = FastAPI(title="Sistema de Incidencias de Producción")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=False,
    allow_methods=["*"],
    allow_headers=["*"],
)

# Seed DB on startup
@app.on_event("startup")
async def startup_event():
    # Run migrations — each in its own connection to avoid aborted transaction state in PostgreSQL
    migrations = [
        "ALTER TABLE reportes_diarios ADD COLUMN tripulacion VARCHAR DEFAULT 'A'",
        "ALTER TABLE lineas ADD COLUMN personas_autorizadas INTEGER DEFAULT 0",
        "ALTER TABLE lineas ADD COLUMN pool_autorizado INTEGER DEFAULT 0",
        "ALTER TABLE lineas ADD COLUMN numerico INTEGER DEFAULT 0",
    ]
    for sql in migrations:
        try:
            with engine.connect() as conn:
                conn.execute(__import__('sqlalchemy').text(sql))
                conn.commit()
        except Exception:
            pass  # Column already exists — ignore

    # Create app_config table if not exists (stores PIN and other settings)
    try:
        with engine.connect() as conn:
            conn.execute(__import__('sqlalchemy').text(
                "CREATE TABLE IF NOT EXISTS app_config (key VARCHAR PRIMARY KEY, value VARCHAR NOT NULL)"
            ))
            conn.commit()
    except Exception:
        pass

    # Create linea_numerico table if not exists
    try:
        with engine.connect() as conn:
            conn.execute(__import__('sqlalchemy').text(
                """CREATE TABLE IF NOT EXISTS linea_numerico (
                    id SERIAL PRIMARY KEY,
                    linea_id INTEGER NOT NULL REFERENCES lineas(id),
                    tripulacion VARCHAR NOT NULL,
                    valor INTEGER NOT NULL DEFAULT 0,
                    CONSTRAINT uq_linea_numerico_trip UNIQUE (linea_id, tripulacion)
                )"""
            ))
            conn.commit()
    except Exception:
        pass

    # Create nomenclatura_mapa table if not exists
    try:
        with engine.connect() as conn:
            conn.execute(__import__('sqlalchemy').text(
                """CREATE TABLE IF NOT EXISTS nomenclatura_mapa (
                    id SERIAL PRIMARY KEY,
                    patron VARCHAR NOT NULL,
                    tripulacion VARCHAR NOT NULL,
                    linea_id INTEGER NOT NULL REFERENCES lineas(id),
                    linea_nombre_ref VARCHAR NOT NULL
                )"""
            ))
            conn.commit()
    except Exception:
        pass

    from seed import seed
    seed()
    from seed_nomenclaturas import seed_nomenclaturas
    seed_nomenclaturas()


# ── Schemas ────────────────────────────────────────────────────────────────────

class IncidenciaIn(BaseModel):
    tipo: str
    cantidad: int
    notas: Optional[str] = None


class ReporteIn(BaseModel):
    linea_id: int
    fecha: date_type
    tripulacion: str = "A"
    lideres_presentes: int
    incidencias: List[IncidenciaIn]


class LineaConfigIn(BaseModel):
    total_lideres: int
    personas_autorizadas: int
    pool_autorizado: int
    numerico: int = 0


class PinIn(BaseModel):
    pin_actual: str
    pin_nuevo: str


# ── Helpers ────────────────────────────────────────────────────────────────────

def _serialize_reporte(reporte, linea):
    return {
        "id": reporte.id,
        "linea_id": reporte.linea_id,
        "linea_nombre": linea.nombre,
        "area_nombre": linea.area.nombre,
        "fecha": reporte.fecha.isoformat(),
        "tripulacion": reporte.tripulacion,
        "lideres_presentes": reporte.lideres_presentes,
        "total_lideres": linea.total_lideres,
        "incidencias": [
            {"id": i.id, "tipo": i.tipo, "cantidad": i.cantidad, "notas": i.notas}
            for i in reporte.incidencias
        ],
        "creado_en": reporte.creado_en.isoformat(),
    }


# ── Endpoints ──────────────────────────────────────────────────────────────────

@app.get("/api/areas")
def get_areas(db: Session = Depends(get_db)):
    areas = crud.get_areas(db)
    return [
        {
            "id": a.id,
            "nombre": a.nombre,
            "lineas": [
                {"id": l.id, "nombre": l.nombre, "total_lideres": l.total_lideres}
                for l in a.lineas
            ],
        }
        for a in areas
    ]


@app.get("/api/lineas")
def get_lineas(area_id: Optional[int] = None, db: Session = Depends(get_db)):
    lineas = crud.get_lineas(db, area_id)
    return [
        {
            "id": l.id,
            "nombre": l.nombre,
            "total_lideres": l.total_lideres,
            "personas_autorizadas": l.personas_autorizadas or 0,
            "pool_autorizado": l.pool_autorizado or 0,
            "numerico": l.numerico or 0,
            "area_id": l.area_id,
            "area_nombre": l.area.nombre,
        }
        for l in lineas
    ]


@app.put("/api/lineas/{linea_id}/config")
def update_linea_config(linea_id: int, data: LineaConfigIn, db: Session = Depends(get_db)):
    linea = crud.get_linea(db, linea_id)
    if not linea:
        raise HTTPException(status_code=404, detail="Línea no encontrada")
    linea.total_lideres = data.total_lideres
    linea.personas_autorizadas = data.personas_autorizadas
    linea.pool_autorizado = data.pool_autorizado
    linea.numerico = data.numerico
    db.commit()
    db.refresh(linea)
    return {
        "id": linea.id,
        "nombre": linea.nombre,
        "total_lideres": linea.total_lideres,
        "personas_autorizadas": linea.personas_autorizadas or 0,
        "pool_autorizado": linea.pool_autorizado or 0,
        "numerico": linea.numerico or 0,
    }


@app.get("/api/lineas/{linea_id}")
def get_linea(linea_id: int, db: Session = Depends(get_db)):
    linea = crud.get_linea(db, linea_id)
    if not linea:
        raise HTTPException(status_code=404, detail="Línea no encontrada")
    return {
        "id": linea.id,
        "nombre": linea.nombre,
        "total_lideres": linea.total_lideres,
        "area_id": linea.area_id,
        "area_nombre": linea.area.nombre,
    }


@app.post("/api/reportes", status_code=201)
def create_reporte(data: ReporteIn, db: Session = Depends(get_db)):
    linea = crud.get_linea(db, data.linea_id)
    if not linea:
        raise HTTPException(status_code=404, detail="Línea no encontrada")
    inc_list = [i.model_dump() for i in data.incidencias]
    reporte = crud.upsert_reporte(db, data.linea_id, data.fecha, data.tripulacion, data.lideres_presentes, inc_list)
    # Re-fetch linea with area after commit
    linea = crud.get_linea(db, data.linea_id)
    reporte = crud.get_reporte(db, data.linea_id, data.fecha, data.tripulacion)
    return _serialize_reporte(reporte, linea)


@app.get("/api/reportes")
def get_reporte(linea_id: int, fecha: date_type, tripulacion: str = "A", db: Session = Depends(get_db)):
    reporte = crud.get_reporte(db, linea_id, fecha, tripulacion)
    if not reporte:
        return None
    linea = crud.get_linea(db, linea_id)
    return _serialize_reporte(reporte, linea)


@app.get("/api/estado-dia")
def get_estado_dia(fecha: date_type, tripulacion: str = "A", db: Session = Depends(get_db)):
    """Devuelve todas las líneas con su estado de carga para una fecha y tripulación dada."""
    lineas = crud.get_lineas(db)
    reportes = {
        r.linea_id: r
        for r in db.query(models.ReporteDiario)
        .options(__import__('sqlalchemy.orm', fromlist=['joinedload']).joinedload(models.ReporteDiario.incidencias))
        .filter(models.ReporteDiario.fecha == fecha, models.ReporteDiario.tripulacion == tripulacion)
        .all()
    }
    resultado = []
    for linea in lineas:
        reporte = reportes.get(linea.id)
        resultado.append({
            "linea_id": linea.id,
            "linea": linea.nombre,
            "area": linea.area.nombre,
            "cargado": reporte is not None,
            "lideres_presentes": reporte.lideres_presentes if reporte else None,
            "total_lideres": linea.total_lideres,
            "total_incidencias": sum(i.cantidad for i in reporte.incidencias) if reporte else None,
            "hora_carga": reporte.creado_en.strftime("%H:%M") if reporte else None,
        })
    total = len(resultado)
    cargados = sum(1 for r in resultado if r["cargado"])
    return {
        "fecha": str(fecha),
        "resumen": {"cargados": cargados, "pendientes": total - cargados, "total": total},
        "lineas": resultado,
    }



@app.get("/api/config/pin-verify")
def verify_pin(pin: str, db: Session = Depends(get_db)):
    """Verifica si el PIN ingresado es correcto."""
    row = db.query(models.AppConfig).filter(models.AppConfig.key == "config_pin").first()
    pin_guardado = row.value if row else "1234"
    return {"ok": pin == pin_guardado}


@app.put("/api/config/pin")
def update_pin(data: PinIn, db: Session = Depends(get_db)):
    """Cambia el PIN. Requiere el PIN actual para autenticarse."""
    row = db.query(models.AppConfig).filter(models.AppConfig.key == "config_pin").first()
    pin_guardado = row.value if row else "1234"
    if data.pin_actual != pin_guardado:
        raise HTTPException(status_code=403, detail="PIN actual incorrecto")
    if len(data.pin_nuevo.strip()) < 4:
        raise HTTPException(status_code=400, detail="El PIN debe tener al menos 4 caracteres")
    if row:
        row.value = data.pin_nuevo.strip()
    else:
        db.add(models.AppConfig(key="config_pin", value=data.pin_nuevo.strip()))
    db.commit()
    return {"ok": True}


# ── Numerico por tripulación (cargado desde Excel) ───────────────────────────

class NumericoMapeoItem(BaseModel):
    texto_excel: str
    tripulacion: str
    conteo: int
    linea_id: Optional[int] = None  # None = ignorar este grupo


class NumericoConfirmarIn(BaseModel):
    mapeos: List[NumericoMapeoItem]


@app.post("/api/numerico/parse-excel")
async def parse_numerico_excel(file: UploadFile = File(...), db: Session = Depends(get_db)):
    """Lee el Excel y devuelve grupos únicos usando la tabla nomenclatura_mapa para auto-mapear."""
    try:
        from openpyxl import load_workbook
    except ImportError:
        raise HTTPException(status_code=500, detail="openpyxl no disponible")

    content = await file.read()
    try:
        wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"No se pudo leer el archivo Excel: {e}")

    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise HTTPException(status_code=400, detail="El archivo está vacío")

    # Detectar fila de encabezados
    header_row = 0
    headers = []
    for i, row in enumerate(rows):
        if any(c is not None for c in row):
            headers = [str(c).strip() if c else '' for c in row]
            header_row = i
            break

    # Buscar columna Supervisory Organization
    col_idx = None
    for i, h in enumerate(headers):
        if 'supervisor' in h.lower() or 'supervisory' in h.lower() or 'org' in h.lower():
            col_idx = i
            break

    if col_idx is None:
        import re
        pat_detect = re.compile(r'(GA\s+.+\s+[ABC]\b|Group Leader GA)', re.IGNORECASE)
        col_counts: dict = {}
        for row in rows[header_row + 1:]:
            for ci, cell in enumerate(row):
                if cell and pat_detect.search(str(cell)):
                    col_counts[ci] = col_counts.get(ci, 0) + 1
        if col_counts:
            col_idx = max(col_counts, key=col_counts.get)

    if col_idx is None:
        raise HTTPException(status_code=400, detail="No se encontró columna Supervisory Organization.")

    # Cargar patrones de la BD para auto-mapear
    mapa_db = db.query(models.NomenclaturaMapa).all()
    # Ordenar de más largo a más corto para evitar falsos positivos de substrings
    mapa_db.sort(key=lambda m: -len(m.patron))

    import re
    EXCLUIR = ['secuenciado', 'qlty', 'quality']
    # Regex para extraer supervisor del paréntesis
    re_supervisor = re.compile(r'\(([^)]+)\)')

    grupos: dict = {}  # key=(texto_celda_grupo, trip, linea_id) → {conteo, supervisores}
    sin_mapeo: dict = {}  # key=(texto_excel, trip) → {conteo, supervisores} — no encontrados en mapa

    for row in rows[header_row + 1:]:
        if col_idx >= len(row) or not row[col_idx]:
            continue
        celda = str(row[col_idx]).strip()
        celda_lower = celda.lower()

        # Filtrar exclusiones
        if any(ex in celda_lower for ex in EXCLUIR):
            continue

        # Extraer supervisor
        m_sup = re_supervisor.search(celda)
        supervisor = m_sup.group(1).strip() if m_sup else ''

        # Buscar patrón en la BD
        matched = None
        for mapa in mapa_db:
            if mapa.patron.lower() in celda_lower:
                matched = mapa
                break

        if matched:
            key = (matched.patron, matched.tripulacion, matched.linea_id)
            if key not in grupos:
                grupos[key] = {"texto_excel": matched.patron, "tripulacion": matched.tripulacion,
                               "linea_id": matched.linea_id, "linea_nombre_ref": matched.linea_nombre_ref,
                               "conteo": 0, "supervisores": set()}
            grupos[key]["conteo"] += 1
            if supervisor:
                grupos[key]["supervisores"].add(supervisor)
        else:
            # No encontrado en mapa — extraer texto para mostrar al BM
            # Intentar detectar tripulación manualmente
            trip_detect = re.search(r'\bGroup Leader GA ([ABC])\b|\bGA\s+.+?\s+([ABC])\s*(?:\(|$)', celda, re.IGNORECASE)
            trip = 'A'
            if trip_detect:
                trip = (trip_detect.group(1) or trip_detect.group(2) or 'A').upper()
            # Extraer texto de línea
            nombre_m = re.search(r'(?:Group Leader )?GA\s+(?:[ABC]\s+)?(.+?)\s+[ABC]\s*(?:\(|$)', celda, re.IGNORECASE)
            texto = nombre_m.group(1).strip() if nombre_m else celda[:60]
            key2 = (texto, trip)
            if key2 not in sin_mapeo:
                sin_mapeo[key2] = {"texto_excel": texto, "tripulacion": trip,
                                   "linea_id": None, "linea_nombre_ref": None,
                                   "conteo": 0, "supervisores": set()}
            sin_mapeo[key2]["conteo"] += 1
            if supervisor:
                sin_mapeo[key2]["supervisores"].add(supervisor)

    resultado = []
    for v in grupos.values():
        resultado.append({**v, "supervisores": sorted(v["supervisores"]), "auto_mapeado": True})
    for v in sin_mapeo.values():
        resultado.append({**v, "supervisores": sorted(v["supervisores"]), "auto_mapeado": False})

    resultado.sort(key=lambda x: (x["tripulacion"], x.get("linea_nombre_ref") or x["texto_excel"]))
    return resultado


@app.get("/api/nomenclatura-mapa")
def get_nomenclatura_mapa(db: Session = Depends(get_db)):
    """Devuelve todos los patrones de nomenclatura configurados."""
    rows = db.query(models.NomenclaturaMapa).order_by(
        models.NomenclaturaMapa.tripulacion, models.NomenclaturaMapa.linea_nombre_ref
    ).all()
    return [{"id": r.id, "patron": r.patron, "tripulacion": r.tripulacion,
             "linea_id": r.linea_id, "linea_nombre_ref": r.linea_nombre_ref} for r in rows]


class NomenclaturaIn(BaseModel):
    patron: str
    tripulacion: str
    linea_id: int
    linea_nombre_ref: str


@app.post("/api/nomenclatura-mapa")
def crear_nomenclatura(data: NomenclaturaIn, db: Session = Depends(get_db)):
    row = models.NomenclaturaMapa(**data.dict())
    db.add(row); db.commit(); db.refresh(row)
    return {"id": row.id, **data.dict()}


@app.put("/api/nomenclatura-mapa/{id}")
def actualizar_nomenclatura(id: int, data: NomenclaturaIn, db: Session = Depends(get_db)):
    row = db.query(models.NomenclaturaMapa).filter(models.NomenclaturaMapa.id == id).first()
    if not row:
        raise HTTPException(status_code=404, detail="No encontrado")
    for k, v in data.dict().items():
        setattr(row, k, v)
    db.commit()
    return {"ok": True}


@app.delete("/api/nomenclatura-mapa/{id}")
def eliminar_nomenclatura(id: int, db: Session = Depends(get_db)):
    row = db.query(models.NomenclaturaMapa).filter(models.NomenclaturaMapa.id == id).first()
    if not row:
        raise HTTPException(status_code=404, detail="No encontrado")
    db.delete(row); db.commit()
    return {"ok": True}


@app.post("/api/numerico/confirmar")
def confirmar_numerico(data: NumericoConfirmarIn, db: Session = Depends(get_db)):
    """Guarda el numérico por línea y tripulación según el mapeo confirmado."""
    guardados = 0
    for item in data.mapeos:
        if item.linea_id is None:
            continue
        existing = db.query(models.LineaNumerico).filter(
            models.LineaNumerico.linea_id == item.linea_id,
            models.LineaNumerico.tripulacion == item.tripulacion,
        ).first()
        if existing:
            existing.valor = item.conteo
        else:
            db.add(models.LineaNumerico(
                linea_id=item.linea_id,
                tripulacion=item.tripulacion,
                valor=item.conteo,
            ))
        guardados += 1
    db.commit()
    return {"ok": True, "guardados": guardados}


@app.get("/api/numerico")
def get_numerico(db: Session = Depends(get_db)):
    """Devuelve todos los numéricos guardados por línea y tripulación."""
    rows = db.query(models.LineaNumerico).all()
    return [{"linea_id": r.linea_id, "tripulacion": r.tripulacion, "valor": r.valor} for r in rows]


@app.get("/api/colaboradores")
def get_colaboradores(
    area_id: Optional[int] = None,
    linea_id: Optional[int] = None,
    tripulacion: Optional[str] = Query(default=None),
    fecha_inicio: Optional[date_type] = Query(default=None),
    fecha_fin: Optional[date_type] = Query(default=None),
    db: Session = Depends(get_db),
):
    """Devuelve incidencias de Incapacidad y Restricción Médica con nombre de colaborador."""
    today = date_type.today()
    if not fecha_inicio:
        fecha_inicio = today.replace(day=1)
    if not fecha_fin:
        fecha_fin = today

    q = (
        db.query(models.Incidencia)
        .options(
            __import__('sqlalchemy.orm', fromlist=['joinedload']).joinedload(models.Incidencia.reporte)
            .joinedload(models.ReporteDiario.linea)
            .joinedload(models.Linea.area)
        )
        .join(models.ReporteDiario)
        .join(models.Linea)
        .filter(
            models.ReporteDiario.fecha >= fecha_inicio,
            models.ReporteDiario.fecha <= fecha_fin,
            models.Incidencia.tipo.in_(["Incapacidad", "Restricción Médica", "Embarazada"]),
            models.Incidencia.notas != None,
            models.Incidencia.notas != "",
        )
    )
    if tripulacion:
        q = q.filter(models.ReporteDiario.tripulacion == tripulacion)
    if linea_id:
        q = q.filter(models.Linea.id == linea_id)
    elif area_id:
        q = q.filter(models.Linea.area_id == area_id)

    resultados = []
    for inc in q.order_by(models.ReporteDiario.fecha.desc()).all():
        resultados.append({
            "fecha": inc.reporte.fecha.isoformat(),
            "area": inc.reporte.linea.area.nombre,
            "linea": inc.reporte.linea.nombre,
            "tripulacion": inc.reporte.tripulacion,
            "tipo": inc.tipo,
            "colaborador": inc.notas,
        })
    return resultados


@app.get("/api/dashboard")
def get_dashboard(
    area_id: Optional[int] = None,
    linea_id: Optional[int] = None,
    tripulacion: Optional[str] = Query(default=None),
    fecha_inicio: Optional[date_type] = Query(default=None),
    fecha_fin: Optional[date_type] = Query(default=None),
    db: Session = Depends(get_db),
):
    today = date_type.today()
    if not fecha_inicio:
        fecha_inicio = today.replace(day=1)
    if not fecha_fin:
        fecha_fin = today
    return crud.get_dashboard_data(db, area_id, linea_id, tripulacion, fecha_inicio, fecha_fin)


@app.get("/api/reporte-ensamble")
def get_reporte_ensamble(
    tripulacion: Optional[str] = Query(default=None),
    fecha_inicio: Optional[date_type] = Query(default=None),
    fecha_fin: Optional[date_type] = Query(default=None),
    db: Session = Depends(get_db),
):
    """Devuelve datos por línea: plantilla autorizada + incidencias capturadas."""
    from sqlalchemy.orm import joinedload as jl
    today = date_type.today()
    if not fecha_inicio:
        fecha_inicio = today
    if not fecha_fin:
        fecha_fin = today

    lineas = db.query(models.Linea).options(jl(models.Linea.area)).all()

    rep_q = (
        db.query(models.ReporteDiario)
        .options(jl(models.ReporteDiario.incidencias), jl(models.ReporteDiario.linea))
        .filter(
            models.ReporteDiario.fecha >= fecha_inicio,
            models.ReporteDiario.fecha <= fecha_fin,
        )
    )
    if tripulacion:
        rep_q = rep_q.filter(models.ReporteDiario.tripulacion == tripulacion)
    reportes = rep_q.all()

    area_order = []
    area_linea_order: dict = {}
    linea_map: dict = {}

    # Cargar numéricos desde linea_numerico (por tripulación), fallback a linea.numerico
    numerico_rows = db.query(models.LineaNumerico).all()
    numerico_dict: dict = {(r.linea_id, r.tripulacion): r.valor for r in numerico_rows}

    for l in lineas:
        an = l.area.nombre
        if an not in area_linea_order:
            area_order.append(an)
            area_linea_order[an] = []
        # Buscar numérico para esta línea+tripulación; si no existe, usar el campo general
        trip_key = tripulacion or ''
        num = numerico_dict.get((l.id, trip_key), None)
        if num is None and trip_key:
            num = l.numerico or 0
        elif num is None:
            # Sin tripulación específica: sumar todos los numéricos de esta línea
            valores = [v for (lid, t), v in numerico_dict.items() if lid == l.id]
            num = sum(valores) if valores else (l.numerico or 0)
        linea_map[l.id] = {
            "id": l.id,
            "nombre": l.nombre,
            "area": an,
            "personas_autorizadas": l.personas_autorizadas or 0,
            "total_lideres": l.total_lideres or 0,
            "pool_autorizado": l.pool_autorizado or 0,
            "numerico": num,
            "total_autorizado": (l.personas_autorizadas or 0) + (l.total_lideres or 0) + (l.pool_autorizado or 0),
            "lets_libres": 0,
            "incidencias": {},
            "total_incidencias": 0,
        }
        area_linea_order[an].append(l.id)

    for rep in reportes:
        lid = rep.linea_id
        if lid in linea_map:
            linea_map[lid]["lets_libres"] += rep.lideres_presentes
            for inc in rep.incidencias:
                t = inc.tipo
                linea_map[lid]["incidencias"][t] = linea_map[lid]["incidencias"].get(t, 0) + inc.cantidad
                linea_map[lid]["total_incidencias"] += inc.cantidad

    return [
        {"area": an, "lineas": [linea_map[lid] for lid in area_linea_order[an]]}
        for an in area_order
    ]


@app.get("/api/export/excel")
def export_excel(
    area_id: Optional[int] = None,
    linea_id: Optional[int] = None,
    tripulacion: Optional[str] = Query(default=None),
    fecha_inicio: Optional[date_type] = Query(default=None),
    fecha_fin: Optional[date_type] = Query(default=None),
    db: Session = Depends(get_db),
):
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise HTTPException(status_code=500, detail="openpyxl no disponible")

    today = date_type.today()
    if not fecha_inicio:
        fecha_inicio = today.replace(day=1)
    if not fecha_fin:
        fecha_fin = today

    data = crud.get_dashboard_data(db, area_id, linea_id, tripulacion, fecha_inicio, fecha_fin)

    wb = Workbook()

    # Estilos
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill_blue = PatternFill("solid", fgColor="1E3A5F")
    header_fill_gray = PatternFill("solid", fgColor="4A4A4A")
    header_fill_green = PatternFill("solid", fgColor="1A6B3C")
    center = Alignment(horizontal="center", vertical="center")
    thin = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )

    def set_header_row(ws, row_num, headers, fill):
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=row_num, column=col, value=h)
            cell.font = header_font
            cell.fill = fill
            cell.alignment = center
            cell.border = thin

    def style_data_row(ws, row_num, num_cols):
        fill = PatternFill("solid", fgColor="F0F4FA") if row_num % 2 == 0 else PatternFill("solid", fgColor="FFFFFF")
        for col in range(1, num_cols + 1):
            c = ws.cell(row=row_num, column=col)
            c.border = thin
            c.fill = fill
            c.alignment = Alignment(horizontal="center", vertical="center")

    def auto_width(ws, min_w=12, max_w=40):
        for col in ws.columns:
            max_len = max((len(str(c.value)) if c.value else 0) for c in col)
            ws.column_dimensions[get_column_letter(col[0].column)].width = min(max(max_len + 2, min_w), max_w)

    # ── Hoja 1: Resumen por área ────────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Resumen por Area"

    ws1.merge_cells("A1:F1")
    title_cell = ws1["A1"]
    title_cell.value = f"Resumen por Area  |  {fecha_inicio} al {fecha_fin}"
    title_cell.font = Font(bold=True, size=13, color="1E3A5F")
    title_cell.alignment = center

    headers1 = ["Area", "Total Incidencias", "Total Lideres", "Lets Libres Prom.", "% Lets Libres", "Tipos de Incidencia"]
    set_header_row(ws1, 2, headers1, header_fill_blue)

    for i, area in enumerate(data["por_area"], start=3):
        desglose_str = ", ".join(f'{d["tipo"]}: {d["cantidad"]}' for d in area.get("desglose", []))
        row = [
            area["area"],
            area["total_incidencias"],
            area["total_lideres"],
            area.get("avg_lideres_presentes") or "-",
            f'{area["pct_lideres_libres"]}%' if area.get("pct_lideres_libres") is not None else "-",
            desglose_str or "Sin incidencias",
        ]
        for col, val in enumerate(row, 1):
            ws1.cell(row=i, column=col, value=val)
        style_data_row(ws1, i, len(headers1))

    auto_width(ws1)

    # ── Hoja 2: Detalle por línea ────────────────────────────────────────────────
    ws2 = wb.create_sheet("Detalle por Linea")

    ws2.merge_cells("A1:G1")
    t2 = ws2["A1"]
    t2.value = f"Detalle por Linea  |  {fecha_inicio} al {fecha_fin}"
    t2.font = Font(bold=True, size=13, color="1A6B3C")
    t2.alignment = center

    headers2 = ["Area", "Linea", "Tripulacion", "Total Incidencias", "Total Lideres", "Lets Libres Prom.", "% Lets Libres"]
    set_header_row(ws2, 2, headers2, header_fill_green)

    for i, linea in enumerate(data["por_linea"], start=3):
        row = [
            linea["area"],
            linea["linea"],
            linea.get("tripulacion", "-"),
            linea["total"],
            linea["total_lideres"],
            linea.get("avg_lideres_presentes") or "-",
            f'{linea["pct_lideres_libres"]}%' if linea.get("pct_lideres_libres") is not None else "-",
        ]
        for col, val in enumerate(row, 1):
            ws2.cell(row=i, column=col, value=val)
        style_data_row(ws2, i, len(headers2))

    auto_width(ws2)

    # ── Hoja 3: Por tipo de incidencia ─────────────────────────────────────────
    ws3 = wb.create_sheet("Por Tipo de Incidencia")

    ws3.merge_cells("A1:C1")
    t3 = ws3["A1"]
    t3.value = f"Por Tipo de Incidencia  |  {fecha_inicio} al {fecha_fin}"
    t3.font = Font(bold=True, size=13, color="4A4A4A")
    t3.alignment = center

    headers3 = ["Tipo de Incidencia", "Total", "% del Total"]
    set_header_row(ws3, 2, headers3, header_fill_gray)

    total_inc = sum(d["total"] for d in data["por_tipo"]) or 1
    for i, tipo in enumerate(data["por_tipo"], start=3):
        pct = round((tipo["total"] / total_inc) * 100, 1)
        row = [tipo["tipo"], tipo["total"], f"{pct}%"]
        for col, val in enumerate(row, 1):
            ws3.cell(row=i, column=col, value=val)
        style_data_row(ws3, i, len(headers3))

    auto_width(ws3)

    # Guardar en buffer y retornar
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"incidencias_{fecha_inicio}_{fecha_fin}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


# ── Serve React frontend (production build) ────────────────────────────────────

STATIC_DIR = pathlib.Path(__file__).parent / "static"

if STATIC_DIR.exists():
    app.mount("/assets", StaticFiles(directory=str(STATIC_DIR / "assets")), name="assets")
    app.mount("/static-files", StaticFiles(directory=str(STATIC_DIR)), name="static-files")

    @app.get("/")
    async def serve_root():
        return FileResponse(str(STATIC_DIR / "index.html"))

    @app.get("/{full_path:path}")
    async def serve_spa(full_path: str):
        file_path = STATIC_DIR / full_path
        if file_path.exists() and file_path.is_file():
            return FileResponse(str(file_path))
        return FileResponse(str(STATIC_DIR / "index.html"))
